from django.conf import settings
from django.core.paginator import Paginator
from django.shortcuts import render, redirect

from .forms import UserRequestForm
from .models import UserRequest
from .utils import save_to_excel
from openpyxl import load_workbook


def home(request):
    if request.method == "POST":
        form = UserRequestForm(data=request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = UserRequestForm()

    context = {
        "form": form
    }
    return render(request, "app/index.html", context)


def elements_view(request):
    query = request.GET
    date_from = query.get('date_from')
    date_to = query.get('date_to')
    phone_number = query.get("phone_number")
    pinfl = query.get("pinfl")
    query_fullname = query.get('fullname')
    passport_series = query.get("passport_series")
    track_number = query.get("track_number")

    if query_fullname:
        elements = UserRequest.objects.filter(fullname__iregex=query_fullname)
    elif phone_number:
        elements = UserRequest.objects.filter(phone_number__iregex=phone_number)
    elif track_number:
        elements = UserRequest.objects.filter(track_number__iregex=track_number)
    elif passport_series:
        elements = UserRequest.objects.filter(passport_number__iregex=passport_series)
    elif pinfl:
        elements = UserRequest.objects.filter(pinfl__iregex=pinfl)
    elif date_from and date_to:
        elements = UserRequest.objects.filter(created_at__gt=date_from, created_at__lt=date_to)
    else:
        elements = UserRequest.objects.all()

    qs = Paginator(elements, 30)
    page = request.GET.get('page')
    elements = qs.get_page(page)

    context = {
        "elements": elements
    }

    return render(request, "app/elements.html", context)


def get_excel(request):
    if request.method == 'POST':
        file = request.FILES.get('file-input')
        wb = load_workbook(file)
        sheet_obj = wb.active
        values = []
        print(sheet_obj.max_row)

        for i in range(1, sheet_obj.max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=1)
            values.append(cell_obj.value)

        result = []
        for value in values:
            elements = UserRequest.objects.filter(track_number=value).values_list(
                'track_number', 'fullname', 'passport_series',
                'passport_number', 'pinfl', 'phone_number').first()

            if elements is None:
                continue

            result.append(elements)

        save_to_excel(
            filename=settings.BASE_DIR / 'app/static/test.xlsx',
            data=result
        )
    else:
        elements = UserRequest.objects.all().values_list('track_number', 'fullname', 'passport_series',
                                                         'passport_number', 'pinfl', 'phone_number')
        save_to_excel(
            filename=settings.BASE_DIR / 'app/static/test.xlsx',
            data=elements
        )

    return render(request, 'app/result.html')
